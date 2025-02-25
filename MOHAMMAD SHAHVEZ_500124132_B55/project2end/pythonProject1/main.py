import os
from openpyxl import Workbook, load_workbook

BASE_DIR = '.'  # Update this to the directory where your Excel files are stored
TOTNUMB_FILE = os.path.join(BASE_DIR, 'totnumb.xlsx')
NAMET_FILE = os.path.join(BASE_DIR, 'namet.xlsx')
ATTENNUM_FILE = os.path.join(BASE_DIR, 'attennum.xlsx')
DEBARLIST_FILE = os.path.join(BASE_DIR, 'debarlist.xlsx')

def create_empty_excel_file(file_path):
    """Create an empty Excel file if it does not exist."""
    if not os.path.exists(file_path):
        wb = Workbook()
        wb.save(file_path)

def add_names():
    try:
        create_empty_excel_file(NAMET_FILE)

        name = input('Enter your name: ')
        sapid = input('Enter your SAPID: ')

        wb = load_workbook(NAMET_FILE)
        sheet = wb.active

        # Append new data to the Excel file
        row = (name, sapid)
        sheet.append(row)

        wb.save(NAMET_FILE)

        create_empty_excel_file(ATTENNUM_FILE)

    except Exception as e:
        print(f'ERROR: {e}')

def search_names():
    try:
        wb = load_workbook(TOTNUMB_FILE)
        sheet = wb.active
        x = sheet['A1'].value  # Assuming 'A1' contains the total number

        wb = load_workbook(NAMET_FILE)
        sheet = wb.active

        a = input('Enter name or SAPID: ')
        count = 0

        for row in sheet.iter_rows(values_only=True):
            count += 1
            if a in row:
                break

        # Retrieve attendance from ATTENNUM_FILE
        wb = load_workbook(ATTENNUM_FILE)
        sheet = wb.active
        attendance = sheet.cell(row=count, column=1).value

        print(f'NAME: {row[0]} \nSAPID: {row[1]} \nATTENDENCE: {attendance} OUT OF {x}')

    except Exception as e:
        print(f'ERROR: {e}')

def mark_att():
    try:
        wb = load_workbook(TOTNUMB_FILE)
        sheet = wb.active
        a = sheet['A1'].value  # Get the total count of attendances

        wb = load_workbook(NAMET_FILE)
        sheet = wb.active

        for row in sheet.iter_rows(values_only=True):
            name = row[0]
            while True:
                decision = input(f'{name}: Press P for Present, A for Absent: ').strip().upper()
                if decision in ('P', 'A'):
                    break
                else:
                    print('Invalid input. Please enter P for Present or A for Absent.')

            if decision == 'P':
                # Update attendance in ATTENNUM_FILE
                wb_atten = load_workbook(ATTENNUM_FILE)
                sheet_atten = wb_atten.active
                current_attendance = sheet_atten.cell(row=row[2], column=1).value
                sheet_atten.cell(row=row[2], column=1, value=current_attendance + 1)
                wb_atten.save(ATTENNUM_FILE)

        # Increment total count in TOTNUMB_FILE
        sheet['A1'] = a + 1
        wb.save(TOTNUMB_FILE)

    except Exception as e:
        print(f'ERROR: {e}')

def debarlist():
    try:
        wb = load_workbook(TOTNUMB_FILE)
        sheet = wb.active
        a = sheet['A1'].value  # Get the total count of attendances

        wb = load_workbook(NAMET_FILE)
        sheet = wb.active

        debarred_students = []

        for row in sheet.iter_rows(values_only=True):
            attendance = row[2]
            attendance_percentage = (attendance / a) * 100

            if attendance_percentage < 75:
                debarred_students.append((row[0], row[1], attendance_percentage))

        if debarred_students:
            wb_debar = Workbook()
            sheet_debar = wb_debar.active
            sheet_debar.append(['Name', 'SAPID', 'Attendance Percentage'])

            for student in debarred_students:
                sheet_debar.append(list(student))

            wb_debar.save(DEBARLIST_FILE)
            print('Debarred students list created successfully.')
        else:
            print('No students are debarred.')

    except Exception as e:
        print(f'ERROR: {e}')

if __name__ == '__main__':
    while True:
        print('PRESS 1 FOR ADDING NAMES')
        print('PRESS 2 FOR MARKING ATTENDANCE')
        print('PRESS 3 FOR VIEWING ATTENDANCE')
        print('PRESS 4 FOR DEBAR LIST')
        print('PRESS 5 FOR EXITING')

        choice = input('Enter your choice: ')

        if choice == '1':
            add_names()
        elif choice == '2':
            mark_att()
        elif choice == '3':
            search_names()
        elif choice == '4':
            debarlist()
        elif choice == '5':
            print('EXITED')
            break
        else:
            print('Please enter a valid option.')
